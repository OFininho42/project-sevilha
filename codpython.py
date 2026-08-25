from pathlib import Path
import re
import openpyxl as opxl


def limpar_nome_pasta(nome):
    # Remove caracteres inválidos no Windows: \ / : * ? " < > |
    return re.sub(r'[\\/*?:"<>|]', '', str(nome)).strip()


class Controle:
    wb_path = r'C:\TEMPORÁRIOS\IMPORTANTE\Meu_Controle.xlsx'
    wb = opxl.load_workbook(wb_path)
    ws = wb['Planilha1']

    codempresa = 'A'
    nomeempresa = 'B'


meu_controle = Controle()

# 1. Armazena os nomes sanitizados
empresas_excel = []

for linha in range(2, meu_controle.ws.max_row + 1):
    val_a = meu_controle.ws[f'{meu_controle.codempresa}{linha}'].value
    val_b = meu_controle.ws[f'{meu_controle.nomeempresa}{linha}'].value

    if val_a is not None or val_b is not None:
        val_a_str = str(val_a).strip() if val_a is not None else ''
        val_b_str = str(val_b).strip() if val_b is not None else ''

        nome_bruto = f'{val_a_str} - {val_b_str}'
        # Sanitiza a string antes de salvar na lista
        nome_limpo = limpar_nome_pasta(nome_bruto)

        if nome_limpo:
            empresas_excel.append(nome_limpo)

# 2. Mapeia a pasta principal e verifica subpastas existentes
caminho_clientes = Path(r'C:\TEMPORÁRIOS\CLIENTES')
caminho_clientes.mkdir(parents=True, exist_ok=True)

pastas_existentes = {p.name for p in caminho_clientes.iterdir() if p.is_dir()}

# 3. Identifica quais empresas ainda não têm pasta
nao_existentes = [
    emp for emp in empresas_excel if emp not in pastas_existentes
]

# 4. Cria as pastas no sistema operacional sem estourar erro de sintaxe
print('=== CRIANDO NOVAS PASTAS ===')
for empresa in nao_existentes:
    caminho_nova_pasta = caminho_clientes / empresa
    caminho_nova_pasta.mkdir(parents=True, exist_ok=True)
    print(f'[CRIADA] {empresa}')

print(
    f'\nSucesso! {len(nao_existentes)} pasta(s) criada(s) em {caminho_clientes}.'
)

#Fellipe